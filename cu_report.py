#! Python3

import openpyxl, datetime, sys

# Table header of purchase report
HEADER_BUY = ['TYP_SUBJEKTU',
 'DIC_DS',
 'NAZEV_DS',
 'DIC_DODAVATELE',
 'NAZEV_DODAVATELE',
 'C_DOKLADU_DPH',
 'NAZEV_VV',
 'NOMENKLATURA',
 'MNOZSTVI',
 'CENA_MJ',
 'DUZP',
 'REZIM_SPD_CLO',
 'D_DOKLADU_SPD',
 'C_NAKL_LISTU',
 'DIC_DOPRAVCE',
 'NAZEV_DOPRAVCE',
 'DRUH_DOPRAVY',
 'RZ_VOZIDLA',
 'RZ_PV',
 'DATUM_NAKLADKY',
 'MISTO_NAKLADKY',
 'DATUM_VYKLADKY',
 'MISTO_VYKLADKY']

# Indexes of columns which have to be copied to purchase report
INDEXES_BUY = [4, 5, 6, 7, 8, 11, 13, 14, 15, 16, 17, 18, 19, 
               20, 22, 23, 24, 25, 26, 27, 28, 29, 30]
 
# Table header of sell report
HEADER_SELL = ['TYP_SUBJEKTU', 
 'DIC_DS',
 'NAZEV_DS',
 'DIC_ODBERATELE',
 'NAZEV_ODBERATELE',
 'C_DOKLADU_DPH',
 'STAV_VV',
 'NAZEV_VV',
 'NOMENKLATURA',
 'MNOZSTVI',
 'CENA_MJ',
 'DUZP',
 'REZIM_SPD_CLO',
 'D_DOKLADU_SPD',
 'C_NAKL_LISTU',
 'DIC_DOPRAVCE',
 'NAZEV_DOPRAVCE',
 'DRUH_DOPRAVY',
 'RZ_VOZIDLA',
 'RZ_PV',
 'DATUM_NAKLADKY',
 'MISTO_NAKLADKY',
 'DATUM_VYKLADKY',
 'MISTO_VYKLADKY']

# Indexes of columns which have to be copied to sell report
INDEXES_SELL = [4, 5, 6, 9, 10, 11, 12, 13, 14, 15, 16, 17, 
                18, 19, 20, 22, 23, 24, 25, 26, 27, 28, 29, 30]


# Update column
def update_column(original_value, updated_value, col, row=5):
    last_row = sheet.max_row + 1
    for i in range(row, last_row):
        value = sheet.cell(row=i, column=col).value
        if value == original_value:
            sheet.cell(row=i, column=col).value = updated_value
            
# Replace specific substring in a cell
def replace_chars(old_char, new_char, column):
    last_row = sheet.max_row + 1
    for i in range(5, last_row):
        if type(sheet.cell(row=i, column=column).value) == str:
            updated_cell = sheet.cell(row=i, column=column).value.replace(old_char, new_char)
            sheet.cell(row=i, column=column).value = updated_cell
  
# Copy column from source to report file          
def copy_column(report, orig_col, new_sheet, new_col, row_from=5):
    orig_sheet = sheet
    last_row = sheet.max_row + 1
    adj = 0
    for i in range(row_from, last_row):
        if orig_sheet.cell(row=i, column=1).value == "NE":
            adj -= 1
        elif orig_sheet.cell(row=i, column=3).value != report:
            adj -= 1
        else:
            new_sheet.cell(row=i-3+adj, column=new_col).value = orig_sheet.cell(row=i, column=orig_col).value
        
# Copy range of columns from source to report file
def copy_range(report, column_numbers, new_sheet):
    index_col = 0
    for n in column_numbers:
        index_col += 1
        copy_column(report, n, new_sheet, index_col)
        
# Create table haeader
def create_header(header, sheet):
    index_col = 0
    for name in header:
        index_col += 1
        sheet.cell(row=1, column=index_col).value = name

# Set format date to DD.MM.YYYY in column
def format_date(sheet, column):
    last_row = sheet.max_row + 1
    for i in range(2, last_row):
        cell = sheet.cell(i, column)
        if type(cell.value) == datetime.datetime:
            dt = cell.value.date().strftime("%d.%m.%Y")
            cell.value = dt

# Set format date in range of columns
def format_date_range(sheet, columns):
    for col in columns:
        format_date(sheet, col)
        
        
# Open source file
period = input("report period(MM_YYYY): ")
print("loading source file...")
try:
    wb = openpyxl.load_workbook("vykaz.xlsx")
except:
    print("""File "vykaz" does not exist!
          Please check the file name""")
sheet = wb.active         

#print("checking source file...")

    
print("updating source file...")

# Update columns
update_column("2 - kartové společnosti", 2, 4)
update_column("1 - distributor PHM", 1, 4)
update_column(None, 26415623, 5)
update_column(None, "W.A.G. payment solutions, a.s.", 6)
update_column(0, None, 16)
update_column(0, None, 19)
replace_chars(" ", "", 25)
replace_chars(" ", "", 26)
replace_chars(",,", ",", 28)


# Sales report
wb_sell = openpyxl.Workbook()
sheet_sell = wb_sell.active
sheet_sell.title = "Prodej"

create_header(HEADER_SELL, sheet_sell)
copy_range("Prodej", INDEXES_SELL, sheet_sell)
format_date_range(sheet_sell, [12, 21, 23])

sell_report_name = "OPHM_P_26415623_" + period + "_R.xlsx"
wb_sell.save(sell_report_name)
print("Sales report created")


# Purchase report
wb_buy = openpyxl.Workbook()
sheet_buy = wb_buy.active
sheet_buy.title = "Nákup" 

create_header(HEADER_BUY, sheet_buy)
copy_range("Nákup", INDEXES_BUY, sheet_buy)
format_date_range(sheet_buy, [11, 20, 22])

buy_report_name = "OPHM_N_26415623_" + period + "_R.xlsx"
wb_buy.save(buy_report_name)
print("Purchase report created")
        
sys.exit



